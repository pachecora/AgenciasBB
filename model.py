import openpyxl
from wtforms import StringField, SubmitField
from wtforms.validators import DataRequired
from flask_wtf import FlaskForm


class Formulario(FlaskForm):
    agencia = StringField('Prefixo', validators=[DataRequired()])
    submit = SubmitField('Processar')


def procuraAgenciasBB(prefixo:str) -> str:

    wb = openpyxl.load_workbook('agenciasBB.xlsx')

    ws = wb.active

    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == prefixo:
            prefixo = ws.cell(row=i, column=2).value
            return prefixo
